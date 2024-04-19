# excel (only support .xlsx)
import  openpyxl

# 新增 excell
# 活頁簿
workbook = openpyxl.Workbook()
# 工作表 #1
sheet = workbook.worksheets[0]
# 以位置寫入
sheet['A1'] = '一年甲班'
list_titles = ['座號','姓名','國文','英文','數學']
# 新增一列
sheet.append(list_titles)
# add item
list_datas = [[1, '大熊', 65, 62, 40],
              [2, '小明', 85, 90, 87],
              [3, '小美', 92, 90, 95]]
for data in list_datas:
    sheet.append(data)
# save
workbook.save('test.xlsx')

# modify excel
# load file
workbook2 = openpyxl.load_workbook('test.xlsx')
sheet = workbook2.worksheets[0]
# sheet.max_row    列數
# sheet.max_column 行數
# sheet.cell(row,column) start from 1
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value, end=' ' )
    print()
sheet['A1'] = '二年甲班'
workbook2.save('test2.xlsx')
# 一年甲班 None None None None
# 座號 姓名 國文 英文 數學
# 1 大熊 65 62 40
# 2 小明 85 90 87
# 3 小美 92 90 95