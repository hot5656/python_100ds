# 儲存所有超市門市資料到 excel
import requests
from bs4 import BeautifulSoup
import openpyxl

def main():
    counties = []
    url = "https://www.ibon.com.tw/retail_inquiry.aspx#gsc.tab=0"
    try:
        resp = requests.get(url)
    except:
        resp = None

    if resp and resp.status_code == 200:
        soup = BeautifulSoup(resp.text, "html.parser")

        datas = soup.select("select#Class1 option")
        for data in datas:
            # print(data.text)
            counties.append(data.text)
        print(counties)

    # 新增 excell
    # 活頁簿
    workbook = openpyxl.Workbook()
    index = 0
    for county in counties:
        get_one_counry(workbook, county, index)
        index += 1

    # remove last sheet
    # last is orgiginal sheet
    last_sheet_name = workbook.sheetnames[-1]
    last_sheet = workbook[last_sheet_name]
    workbook.remove(last_sheet)

    # save
    workbook.save('711stores.xlsx')


def get_one_counry(workbook, county, index):
    # index=0 1st sheet
    sheet = workbook.create_sheet(title=county, index=index)

    ajax_url = 'https://www.ibon.com.tw/retail_inquiry_ajax.aspx'
    payload = {
        'strTargetField': 'COUNTY',
        'strKeyWords': county
    }
    try:
        ajax_resp = requests.post(ajax_url, data=payload)
    except:
        ajax_resp = None

    if ajax_resp and ajax_resp.status_code == 200:
        soup = BeautifulSoup(ajax_resp.text, 'html.parser')
        tables = soup.select("tr")

        for data in tables:
            record = []
            items = data.select("td")
            for item in items:
                record.append(item.text.strip())
                # print(item.text.strip(), end=',' )
            # print()
            sheet.append(record)
    else:
        print("get data error ...")

if __name__ == '__main__':
    main()